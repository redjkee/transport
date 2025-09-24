import pandas as pd
import openpyxl
import re
from pathlib import Path
from datetime import datetime
import os

def detect_month_from_files(files):
    """Автоматически определяет месяц из названий файлов"""
    month_names = {
        '01': 'январь', '02': 'февраль', '03': 'март', '04': 'апрель',
        '05': 'май', '06': 'июнь', '07': 'июль', '08': 'август',
        '09': 'сентябрь', '10': 'октябрь', '11': 'ноябрь', '12': 'декабрь'
    }
    
    # Пробуем найти месяц в названиях файлов
    for file_path in files:
        filename = file_path.name.lower()
        
        # Ищем русские названия месяцев
        for month_num, month_name in month_names.items():
            if month_name in filename:
                return month_name
        
        # Ищем числовое представление месяца
        month_match = re.search(r'(\d{2})\.?xlsx', filename)
        if month_match:
            month_num = month_match.group(1)
            if month_num in month_names:
                return month_names[month_num]
    
    # Если месяц не найден в файлах, используем текущий
    current_month = datetime.now().month
    return month_names.get(str(current_month).zfill(2), "текущий_месяц")

def find_table_structure(ws):
    """Находит структуру таблицы по ключевым заголовкам"""
    print("🔍 Ищу структуру таблицы...")
    
    headers_positions = {}
    
    for row in ws.iter_rows():
        for cell in row:
            if cell.value:
                cell_value = str(cell.value).strip()
                
                if "Товары (работы, услуги)" in cell_value:
                    headers_positions['description'] = (cell.row, cell.column)
                    print(f"📋 Найден заголовок 'Товары' в строке {cell.row}, столбце {cell.column}")
                elif "Сумма" in cell_value and cell_value != "Сумма с НДС":
                    headers_positions['amount'] = (cell.row, cell.column)
                    print(f"💰 Найден заголовок 'Сумма' в строке {cell.row}, столбце {cell.column}")
                elif "№" == cell_value and cell.column < 10:
                    headers_positions['number'] = (cell.row, cell.column)
                    print(f"🔢 Найден заголовок '№' в строке {cell.row}, столбце {cell.column}")
                elif "Кол-во" in cell_value:
                    headers_positions['quantity'] = (cell.row, cell.column)
                    print(f"📦 Найден заголовок 'Кол-во' в строке {cell.row}, столбце {cell.column}")
                elif "Ед." in cell_value:
                    headers_positions['unit'] = (cell.row, cell.column)
                    print(f"📏 Найден заголовок 'Ед.' в строке {cell.row}, столбце {cell.column}")
                elif "Цена" in cell_value:
                    headers_positions['price'] = (cell.row, cell.column)
                    print(f"🏷️ Найден заголовок 'Цена' в строке {cell.row}, столбце {cell.column}")
    
    return headers_positions

def extract_data_from_description(description):
    """Извлекает дату, маршрут, гос. номер и фамилию водителя из описания"""
    description_str = str(description)
    
    # Маршрут (все до первой запятой)
    route = description_str.split(',')[0].strip()
    
    # Дата из текста (формат "от 06.09.25")
    date_match = re.search(r'от\s+(\d{2}\.\d{2}\.\d{2})', description_str)
    date_str = date_match.group(1) if date_match else "Дата не найдена"
    
    # Гос. номер - ищем 3 цифры подряд
    plate_match = re.search(r'(\d{3})', description_str)
    car_plate = plate_match.group(1) if plate_match else "Неизвестно"
    
    # Фамилия водителя
    driver_match = re.search(r',\s*([А-Я][а-я]+)\s+[А-Я]\.[А-Я]\.', description_str)
    if driver_match:
        driver_name = driver_match.group(1)
    else:
        alt_driver_match = re.search(r',\s*([А-Я][а-я]+)', description_str)
        driver_name = alt_driver_match.group(1) if alt_driver_match else "Фамилия не найдена"
    
    return route, date_str, car_plate, driver_name

def parse_invoice_file(file_path):
    """Парсит один файл счета и возвращает данные"""
    try:
        print(f"\n🔍 Обрабатываю файл: {file_path.name}")
        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb.active
        
        headers = find_table_structure(ws)
        
        if not headers.get('description') or not headers.get('amount'):
            print("⚠️ Не найдена полная структура таблицы")
            return []
        
        header_row = max(h[0] for h in headers.values())
        description_col = headers['description'][1]
        amount_col = headers['amount'][1]
        
        print(f"📊 Структура таблицы определена:")
        print(f"   - Строка данных: {header_row + 1}")
        print(f"   - Столбец описания: {description_col}")
        print(f"   - Столбец суммы: {amount_col}")
        
        parsed_data = []
        row_num = header_row + 1
        processed_count = 0
        current_empty_rows = 0
        max_empty_rows = 5
        
        print(f"🔎 Читаю данные таблицы...")
        
        while current_empty_rows < max_empty_rows:
            description_cell = ws.cell(row=row_num, column=description_col)
            description = description_cell.value
            
            if not description:
                current_empty_rows += 1
                row_num += 1
                continue
                
            current_empty_rows = 0
            description_str = str(description)
            
            if any(word in description_str.lower() for word in ['итого', 'всего', 'итог', 'сумма']):
                row_num += 1
                continue
            
            amount_cell = ws.cell(row=row_num, column=amount_col)
            amount = amount_cell.value
            
            if amount is not None:
                try:
                    if isinstance(amount, str) and any(char.isalpha() for char in amount.replace(' ', '').replace(',', '.')):
                        row_num += 1
                        continue
                    
                    amount_str = str(amount).replace(' ', '').replace(',', '.')
                    amount_value = float(amount_str)
                    
                    route, date_str, car_plate, driver_name = extract_data_from_description(description_str)
                    
                    if car_plate != "Неизвестно" and amount_value > 0:
                        parsed_data.append({
                            'Дата': date_str,
                            'Маршрут': route,
                            'Стоимость': amount_value,
                            'Гос_номер': car_plate,
                            'Водитель': driver_name,
                            'Источник': file_path.name,
                            'Строка': row_num
                        })
                        processed_count += 1
                        print(f"✅ Строка {row_num}: {date_str} | {route[:20]}... | {car_plate} | {driver_name} | {amount_value:,.0f} руб.")
                    
                except (ValueError, TypeError):
                    pass
            
            row_num += 1
            
            if row_num > header_row + 1000:
                print("⚠️ Достигнуто ограничение в 1000 строк")
                break
        
        print(f"📊 Обработано записей: {processed_count}")
        return parsed_data
        
    except Exception as e:
        print(f"❌ Ошибка при обработке {file_path}: {e}")
        return []

def create_report(data, columns, output_path, sheet_name="Отчет"):
    """Создает отчет без жирного шрифта"""
    if data.empty:
        return
    
    # Добавляем итоговую строку (без жирного шрифта)
    total_row = pd.DataFrame([{col: '' for col in columns}])
    total_row['Стоимость'] = data['Стоимость'].sum()
    total_row[columns[0]] = 'ИТОГО:'
    
    final_df = pd.concat([data, total_row], ignore_index=True)
    
    # Сохраняем в Excel
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        final_df.to_excel(writer, sheet_name=sheet_name, index=False)
        worksheet = writer.sheets[sheet_name]
        
        # Настраиваем ширину столбцов
        column_widths = {'Дата': 12, 'Маршрут': 30, 'Водитель': 15, 'Гос_номер': 12, 'Стоимость': 15}
        
        for i, col in enumerate(columns, 1):
            col_letter = chr(64 + i)
            worksheet.column_dimensions[col_letter].width = column_widths.get(col, 15)

def data_quality_check(df):
    """Проверяет качество данных и выдает подробный отчет"""
    print(f"\n🔍 ПРОВЕРКА КАЧЕСТВА ДАННЫХ...")
    
    issues = []
    detailed_info = []
    
    # Проверка на дубликаты
    duplicates = df.duplicated().sum()
    if duplicates > 0:
        issues.append(f"⚠️ Найдено {duplicates} дубликатов")
        dup_data = df[df.duplicated()]
        for _, row in dup_data.iterrows():
            detailed_info.append(f"   Дубликат: {row['Дата']} | {row['Маршрут'][:30]}... | {row['Водитель']} | {row['Стоимость']:,.0f} руб.")
    
    # Проверка на пропущенные даты
    missing_dates = df[df['Дата'] == 'Дата не найдена']
    if len(missing_dates) > 0:
        issues.append(f"⚠️ В {len(missing_dates)} записях не найдена дата")
        for _, row in missing_dates.iterrows():
            detailed_info.append(f"   Нет даты: {row['Маршрут'][:30]}... | {row['Водитель']} | {row['Стоимость']:,.0f} руб.")
    
    # Проверка на нераспознанные водителей
    unknown_drivers = df[df['Водитель'] == 'Фамилия не найдена']
    if len(unknown_drivers) > 0:
        issues.append(f"⚠️ В {len(unknown_drivers)} записях не найдена фамилия водителя")
        for _, row in unknown_drivers.iterrows():
            detailed_info.append(f"   Нет водителя: {row['Дата']} | {row['Маршрут'][:30]}... | {row['Гос_номер']} | {row['Стоимость']:,.0f} руб.")
    
    # Проверка на нераспознанные автомобили
    unknown_cars = df[df['Гос_номер'] == 'Неизвестно']
    if len(unknown_cars) > 0:
        issues.append(f"⚠️ В {len(unknown_cars)} записях не найден гос номер")
        for _, row in unknown_cars.iterrows():
            detailed_info.append(f"   Нет номера: {row['Дата']} | {row['Маршрут'][:30]}... | {row['Водитель']} | {row['Стоимость']:,.0f} руб.")
    
    # Проверка на аномальные суммы
    if len(df) > 0:
        avg_amount = df['Стоимость'].mean()
        std_amount = df['Стоимость'].std()
        anomalies = df[df['Стоимость'] > avg_amount + 2 * std_amount]
        if len(anomalies) > 0:
            issues.append(f"⚠️ Найдено {len(anomalies)} аномальных сумм (выше {avg_amount + 2 * std_amount:,.0f} руб.)")
            for _, row in anomalies.iterrows():
                detailed_info.append(f"   Аномальная сумма: {row['Дата']} | {row['Маршрут'][:30]}... | {row['Водитель']} | {row['Стоимость']:,.0f} руб. (среднее: {avg_amount:,.0f} руб.)")
    
    # Выводим результаты проверки
    if issues:
        print("❌ Проблемы с качеством данных:")
        for issue in issues:
            print(f"   {issue}")
        
        print(f"\n📋 Подробный список проблем:")
        for info in detailed_info:
            print(info)
    else:
        print("✅ Качество данных отличное! Проблем не найдено.")
    
    return issues, detailed_info

def main():
    # ПУТИ
    input_folder = Path("C:/папка_с_отчетами")
    base_output_folder = Path("C:/folder")
    
    # Создаем папки для результатов
    auto_folder = base_output_folder / "авто"
    driver_folder = base_output_folder / "водитель"
    
    auto_folder.mkdir(parents=True, exist_ok=True)
    driver_folder.mkdir(parents=True, exist_ok=True)
    
    # Проверяем существование папки с отчетами
    if not input_folder.exists():
        print(f"❌ Папка с отчетами не найдена: {input_folder}")
        return
    
    # Проверяем файлы в папке
    files = list(input_folder.glob("*.xlsx"))
    print(f"📁 Найдено файлов: {len(files)}")
    
    if not files:
        print("❌ Нет Excel-файлов для обработки")
        return
    
    # Автоматически определяем месяц
    month_name = detect_month_from_files(files)
    print(f"📅 Автоматически определен месяц: {month_name}")
    
    # Собираем все данные
    all_data = []
    
    print("\n🔍 Начинаю обработку файлов...")
    for file_path in files:
        if '~' in file_path.name:
            continue
        file_data = parse_invoice_file(file_path)
        all_data.extend(file_data)
        print("=" * 80)
    
    if not all_data:
        print("❌ Не найдено данных для обработки")
        return
    
    # Создаем DataFrame
    df = pd.DataFrame(all_data)
    
    # Проверка качества данных (с подробным выводом)
    data_quality_check(df)
    
    print(f"\n🎉 ОБРАБОТКА ЗАВЕРШЕНА!")
    print(f"📈 Найдено записей: {len(all_data)}")
    print(f"🚗 Уникальные гос. номера: {', '.join([x for x in df['Гос_номер'].unique() if x != 'Неизвестно'])}")
    print(f"👤 Уникальные водители: {', '.join([x for x in df['Водитель'].unique() if x != 'Фамилия не найдена'])}")
    
    # СОЗДАЕМ ОТЧЕТЫ ПО АВТОМОБИЛЯМ
    print(f"\n🚗 СОЗДАЮ ОТЧЕТЫ ПО АВТОМОБИЛЯМ...")
    for car_plate, car_data in df.groupby('Гос_номер'):
        if car_plate == "Неизвестно":
            continue
            
        car_data_sorted = car_data.sort_values('Дата')
        report_data = car_data_sorted[['Дата', 'Маршрут', 'Водитель', 'Стоимость']].copy()
        
        filename = f"{car_plate} {month_name}.xlsx"
        filepath = auto_folder / filename
        
        create_report(report_data, ['Дата', 'Маршрут', 'Водитель', 'Стоимость'], filepath)
        
        total_amount = report_data['Стоимость'].sum()
        print(f"💾 Авто отчет: {filename} (записей: {len(report_data)}, сумма: {total_amount:,.0f} руб.)")
    
    # СОЗДАЕМ ОТЧЕТЫ ПО ВОДИТЕЛЯМ
    print(f"\n👤 СОЗДАЮ ОТЧЕТЫ ПО ВОДИТЕЛЯМ...")
    for driver_name, driver_data in df.groupby('Водитель'):
        if driver_name == "Фамилия не найдена":
            continue
            
        driver_data_sorted = driver_data.sort_values('Дата')
        report_data = driver_data_sorted[['Дата', 'Маршрут', 'Гос_номер', 'Стоимость']].copy()
        
        safe_driver_name = re.sub(r'[<>:"/\\|?*]', '_', driver_name)
        filename = f"{safe_driver_name} {month_name}.xlsx"
        filepath = driver_folder / filename
        
        create_report(report_data, ['Дата', 'Маршрут', 'Гос_номер', 'Стоимость'], filepath)
        
        total_amount = report_data['Стоимость'].sum()
        car_plates = ', '.join(driver_data_sorted['Гос_номер'].unique())
        print(f"💾 Водитель отчет: {filename} (записей: {len(report_data)}, сумма: {total_amount:,.0f} руб.)")
    
    # ФИНАЛЬНАЯ СТАТИСТИКА
    print(f"\n📊 ФИНАЛЬНАЯ СТАТИСТИКА:")
    print(f"📁 Отчеты по автомобилям: {auto_folder}")
    print(f"📁 Отчеты по водителям: {driver_folder}")
    print(f"🚗 Обработано автомобилей: {df['Гос_номер'].nunique() - 1}")
    print(f"👤 Обработано водителей: {df['Водитель'].nunique() - 1}")
    print(f"💰 Общая сумма: {df['Стоимость'].sum():,.0f} руб.")

if __name__ == "__main__":
    main()
    #test_2